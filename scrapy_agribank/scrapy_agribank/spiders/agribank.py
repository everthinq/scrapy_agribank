import json
import scrapy
from openpyxl import load_workbook


class agribank(scrapy.Spider):
    name = 'agribank'

    def start_requests(self):
        region_codes = ['adg',
                        'altai',
                        'amur',
                        'archangelic',
                        'bashkiria',
                        'belgorod',
                        'bryansk',
                        'buryatia',
                        'vladimir',
                        'volgograd',
                        'voronezh',
                        'dagestan',
                        'ivanovo',
                        'ingushetia',
                        'irkutsk',
                        'kbr',
                        'kaliningrad',
                        'kaluga',
                        'kamchatka',
                        'kemerovo',
                        'kirov',
                        'komi',
                        'kostroma',
                        'krasnodar',
                        'krasnoyarsk',
                        'kursk',
                        'lipetsk',
                        'mariel',
                        'mordovia',
                        'moscow',
                        'nnovgorod',
                        'novgorod',
                        'novosibirsk',
                        'omsk',
                        'orenburg',
                        'orel',
                        'penza',
                        'perm',
                        'primorye',
                        'pskov',
                        'rostov',
                        'ryazan',
                        'samara',
                        'spb',
                        'saratov',
                        'sakhalin',
                        'sverdlovsk',
                        'smolensk',
                        'stavropol',
                        'tambov',
                        'tatarstan',
                        'tver',
                        'tomsk',
                        'tyva',
                        'tula',
                        'tumen',
                        'udmurtia',
                        'ulyanovsk',
                        'khabarovsk',
                        'ckb',
                        'chelyabinsk',
                        'chechnya',
                        'chita',
                        'chuvashia',
                        'yakutsk',
                        'yaroslavl']
        self.cell_value = '2'
        self.workbook = load_workbook('5. JSC Russian Agricultural Bank Russia.xlsx')
        self.worksheet = self.workbook[self.workbook.sheetnames[0]]

        for region in region_codes:
            post_body = 'branchCode=' + region + '&locality=%D0%92%D1%81%D0%B5&type=offices.list&query='
            yield scrapy.Request(
                method='POST',
                url='https://www.rshb.ru/ajax/get-data.php',
                headers={'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8'},
                body=post_body,
                callback=self.parse_json)
            # break

    def parse_json(self, response):
        if (response.text):
            JSON = json.loads(response.text)

            for branch_id in JSON['officeItems']:
                branch_name = JSON['officeItems'][branch_id]['name']
                address = JSON['officeItems'][branch_id]['address']
                lat = JSON['officeItems'][branch_id]['location_lat']
                lng = JSON['officeItems'][branch_id]['location_lng']

                print('Writing --', 'type:', JSON['officeItems'][branch_id]['type'], 'name:', branch_name, 'address:',
                      address, 'lat:', lat, 'lng:', lng)

                self.cell_value = str(self.cell_value)
                self.worksheet['B' + self.cell_value] = 'JSC Russian Agricultural Bank Russia'
                self.worksheet['C' + self.cell_value] = branch_name
                self.worksheet['D' + self.cell_value] = address
                self.worksheet['G' + self.cell_value] = 'Russia'
                self.worksheet['H' + self.cell_value] = 'RU'
                self.worksheet['M' + self.cell_value] = lat
                self.worksheet['N' + self.cell_value] = lng
                self.worksheet['O' + self.cell_value] = 'Address'
                self.worksheet['R' + self.cell_value] = 'Bank website'
                self.cell_value = int(self.cell_value) + 1

        self.workbook.save('5. JSC Russian Agricultural Bank Russia.xlsx')
